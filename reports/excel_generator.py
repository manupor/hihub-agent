"""
Excel Generator - Genera hojas de cálculo con información de clientes
"""
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_excel(data, output_path):
    """
    Genera un archivo Excel con la información del cliente
    
    Args:
        data: Diccionario con información formateada
        output_path: Ruta donde guardar el archivo Excel
    
    Returns:
        str: Ruta del archivo generado
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Clientes HiHub"
    
    # Definir columnas
    headers = [
        "Fecha", "Cliente", "Teléfono", "Empresa", "Producto",
        "Especificaciones", "Cantidad", "Ubicación Cliente", "Ubicación Envío",
        "Transporte", "Estado", "Precio Estimado", "Fecha Entrega",
        "Forma de Pago", "Notas", "Adjuntos"
    ]
    
    # Estilos
    header_fill = PatternFill(start_color="F7941D", end_color="F7941D", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Escribir encabezados
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Escribir datos
    data["fecha"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    row_data = [
        data.get("fecha", ""),
        data.get("cliente", ""),
        data.get("telefono", ""),
        data.get("empresa", ""),
        data.get("producto", ""),
        data.get("especificaciones", ""),
        data.get("cantidad", ""),
        data.get("ubicacion_cliente", ""),
        data.get("ubicacion_envio", ""),
        data.get("transporte", "Por Definir"),
        data.get("estado", "Consulta Inicial"),
        data.get("precio_estimado", ""),
        data.get("fecha_entrega", ""),
        data.get("forma_pago", ""),
        data.get("notas", ""),
        data.get("adjuntos", "")
    ]
    
    for col_num, value in enumerate(row_data, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.value = value
        cell.border = border
        cell.alignment = Alignment(vertical='top', wrap_text=True)
    
    # Ajustar anchos de columna
    column_widths = {
        'A': 18, 'B': 20, 'C': 15, 'D': 20, 'E': 25,
        'F': 30, 'G': 12, 'H': 18, 'I': 18, 'J': 15,
        'K': 18, 'L': 15, 'M': 15, 'N': 15, 'O': 30, 'P': 20
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Ajustar altura de fila de datos
    ws.row_dimensions[2].height = 60
    
    # Guardar archivo
    wb.save(output_path)
    return output_path
